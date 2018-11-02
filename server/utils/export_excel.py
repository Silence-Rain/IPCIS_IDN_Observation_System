import io
import openpyxl

def export_excel(filename):
	file = "%s.xlsx" % filename
	wb=openpyxl.Workbook()
	wb.save(filename=file)
	wb = openpyxl.load_workbook(file)
	sheet = wb.active
	sheet.title = "search_data"

	value = []
	with io.open(filename, "r", encoding="utf8") as f:
		# 原始数据每行是一个对象
		for line in f.readlines():
			content = eval(line)
			value.append(list(content.values()))
		# # 原始数据是一个str(数组)
		# content = eval(f.readline())
		# for item in content.values():
		# 	value.append(list(item.values()))

	for i in range(0, len(value)):
	    for j in range(0, len(value[i])):
	        sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
	wb.save(file)